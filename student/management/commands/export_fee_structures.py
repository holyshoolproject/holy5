
# app_name/management/commands/export_fee_structures.py
from django.core.management.base import BaseCommand
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from fees.models import FeeStructure  # adjust app_name

class Command(BaseCommand):
    help = "Export FeeStructure data to an Excel file"

    def add_arguments(self, parser):
        parser.add_argument(
            "--outfile",
            type=str,
            default=f"fee_structures_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            help="Output Excel filename (default includes timestamp).",
        )

    def handle(self, *args, **options):
        outfile = options["outfile"]

        qs = (
            FeeStructure.objects
            .select_related("academic_year", "grade_class", "term")
            .order_by("academic_year_id", "grade_class_id", "term_id")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Fee Structures"

        # Header row
        headers = [
            "ID",
            "Academic Year",
            "Grade/Class",
            "Term",
            "Amount",
        ]
        ws.append(headers)

        # Data rows
        for fs in qs:
            ws.append([
                fs.id,
                # Adjust field accessors based on your models (e.g., fs.academic_year.name)
                getattr(fs.academic_year, "name", str(fs.academic_year)),
                getattr(fs.grade_class, "name", str(fs.grade_class)),
                getattr(fs.term, "name", str(fs.term)),
                float(fs.amount),
            ])

        # Autosize columns
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(header)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                cell_val = row[0].value
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        wb.save(outfile)
        self.stdout.write(self.style.SUCCESS(f"Exported {qs.count()} fee structure(s) to {outfile}"))
