
# app_name/management/commands/export_student_fee_records.py
from django.core.management.base import BaseCommand
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from fees.models import StudentFeeRecord  # adjust app_name

class Command(BaseCommand):
    help = "Export StudentFeeRecord data to an Excel file"

    def add_arguments(self, parser):
        parser.add_argument(
            "--outfile",
            type=str,
            default=f"student_fee_records_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            help="Output Excel filename (default includes timestamp).",
        )

    def handle(self, *args, **options):
        outfile = options["outfile"]

        qs = StudentFeeRecord.objects.select_related("student", "fee_structure").order_by("date_created")

        wb = Workbook()
        ws = wb.active
        ws.title = "Student Fee Records"

        # Header row
        headers = [
            "Student ID",
            "Student Name",
            "Fee Structure",
            "Amount Paid",
            "Balance",
            "Is Fully Paid",
            "Date Created",
        ]
        ws.append(headers)

        # Data rows
        for r in qs:
            ws.append([
                r.student.id if r.student_id else "",
                getattr(r.student, "full_name", str(r.student)),  # adjust if you have a specific name field
                str(r.fee_structure),  # or r.fee_structure.name if exists
                float(r.amount_paid),
                float(r.balance),
                "Yes" if r.is_fully_paid else "No",
                timezone.localtime(r.date_created).strftime("%Y-%m-%d %H:%M:%S"),
            ])

        # Autosize columns
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(header)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                cell_val = row[0].value
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        wb.save(outfile)
        self.stdout.write(self.style.SUCCESS(f"Exported {qs.count()} records to {outfile}"))
